import openpyxl
import random
from openpyxl import Workbook, load_workbook
# from kobert_transformers import get_tokenizer
from kogpt2_transformers import get_kogpt2_tokenizer
import json
import os
import re
######################## 데이터 전처리 #################################

# ######################### 공감형 대화셋 호출 후 저장#######################################
# def tweet_dialog_dataset(json_files_path, output_file_path):       
#     with open(output_file_path, 'w', encoding='utf-8') as output_file:
#         for filename in os.listdir(json_files_path):
#             if filename.endswith(".json"):
#                 file_path = os.path.join(json_files_path, filename)
#                 with open(file_path, 'r', encoding='utf-8') as json_file:
#                     print(f"Processing file: {filename}")
#                     data = json.load(json_file)
#                     utterances = data.get('utterances', '')
#                     if utterances:
#                         print(f"Writing utterance: {utterances}") 
#                         utterances_str = "\n".join(str(utterance) for utterance in utterances)
#                         output_file.write(utterances_str + "\n\n\n")
#                     else:
#                         print(f"No utterance in file: {filename}") 

######################## Wellness 대화셋 질문부분을 추출 후 저장#######################################
# def wellness_question_data():
#   root_path = "C:\\sqlite\\mysql\\code\\AI\\FINAL_project\\WellnessConversation-LanguageModel-master\\WellnessConversation-LanguageModel-master\\dataloader"
#   wellness_file = root_path + "\\wellness_dialog_dataset.xlsx"
#   wellness_q_output = root_path + "\\wellness_dialog_question.txt"

#   f = open(wellness_q_output, 'w')
#   wb = load_workbook(filename=wellness_file)
#   ws = wb[wb.sheetnames[0]]
#   # print(sheet)
#   for row in ws.iter_rows():
#     value1 = row[0].value if row[0].value is not None else ''
#     value2 = row[1].value if row[1].value is not None else ''
    
#     print(f"Writing: {value1}    {value2}")
#     f.write(value1 + "    " + value2 + "\n")

#   f.close()


########################  Wellness 대화 데이터 셋 답변 부분 추출 후 저장#######################################

# def wellness_answer_data():
#   root_path = "C:\\sqlite\\mysql\\code\\AI\\FINAL_project\\WellnessConversation-LanguageModel-master\\WellnessConversation-LanguageModel-master\\dataloader"
#   wellness_file = root_path + "\\wellness_dialog_dataset.xlsx"
#   wellness_a_output = root_path + "\\wellness_dialog_answer.txt"

#   f = open(wellness_a_output, 'w')
#   wb = load_workbook(filename=wellness_file)
#   ws = wb[wb.sheetnames[0]]

#   for row in ws.iter_rows():
#     if row[2].value == None:
#       continue
#     else:
#       f.write(row[0].value + "    " + row[2].value + "\n")
#   f.close()


######################## 챗봇 대화 데이터 셋 카테고리 확인, 저장#######################################
# def category_data():
#   root_path = "C:\\sqlite\\mysql\\code\\AI\\FINAL_project\\dialogLM\\dialogLM\\data"
#   data_path = root_path + "\\chatbot_wellness_data.txt"
#   c_output = root_path + "\\chatbot_wellness_category.txt"

#   i_f = open(data_path, 'r', encoding='utf-8')
#   o_f = open(c_output, 'w')

#   category_count = 0
#   flag = True

#   cate_dict = []
#   i_lines = i_f.readlines()
#   for i, data in enumerate(i_lines):
#     tmp = data.split(',')
#     print(tmp)

#     try:  # try-except 블록 추가
#       a = tmp[1][:-1]
#       q = tmp[0]
#       if a not in cate_dict:
#         cate_dict.append(a)
#         o_f.write(a.strip() + "    " + str(category_count) + "\n")
#         category_count += 1
#     except IndexError:
#       print(f"Error {i}: {tmp}") 
#   o_f.close()
#   i_f.close()


######################## Wellness 대화 ,카테고리 맵핑 후 저장#######################################

# def wellness_text_classification_data():
#     root_path = "C:\\sqlite\\mysql\\code\\AI\\FINAL_project\\WellnessConversation-LanguageModel-master\\WellnessConversation-LanguageModel-master\\dataloader"
#     wellness_category_file = root_path + "\\wellness_dialog_category.txt"
#     wellness_question_file = root_path + "\\wellness_dialog_question.txt"
#     wellness_text_classification_file = root_path + "\\wellness_dialog_for_text_classification.txt"

  
#     cate_file = open(wellness_category_file, 'r', encoding='utf8')
#     ques_file = open(wellness_question_file, 'r')
#     text_classfi_file = open(wellness_text_classification_file, 'w', encoding='utf8')  

#     cate_dict = {}
#     category_lines = cate_file.readlines()
#     for line_data in category_lines:
#         data = line_data.split('    ')
#         cate_dict[data[0]] = data[1][:-1]

#     ques_dict = {}
#     ques_lines = ques_file.readlines()
#     for line_data in ques_lines:
#         data = line_data.split('    ')
#         if data[0] in cate_dict: 
#             text_classfi_file.write(data[1][:-1] + "    " + cate_dict[data[0]] + "\n")

#     cate_file.close()
#     ques_file.close()
#     text_classfi_file.close()

##################################전처리 수정본################################################################

# def wellness_text_classification_data():
#     root_path = "C:\\sqlite\\mysql\\code\\AI\\FINAL_project\\dialogLM\\dialogLM\\data"
#     wellness_category_file = root_path + "\\wellness_dialog_category.txt"
#     wellness_question_file = root_path + "\\wellness_dialog_question.txt"
#     wellness_text_classification_file = root_path + "\\wellness_dialog_for_text_classification_train.txt"

#     with open(wellness_category_file, 'r', encoding='utf8') as cate_file, \
#          open(wellness_question_file, 'r') as ques_file, \
#          open(wellness_text_classification_file, 'w', encoding='utf8') as text_classfi_file:
        
#         cate_dict = {}
#         category_lines = cate_file.readlines()
#         for line_data in category_lines:
#             data = line_data.split('    ')
#             if len(data) >= 2:
#                 cate_dict[data[0]] = data[1].strip()
#         print("Category Dictionary: ", cate_dict) 
#         ques_dict = {}
#         ques_lines = ques_file.readlines()
#         for line_data in ques_lines:
#             data = line_data.split('    ')
#             if len(data) >= 2 and data[0] in cate_dict:
#                 text_classfi_file.write(data[1].strip() + "    " + cate_dict[data[0]] + "\n")
#             elif len(data) >= 2:  
#                 print("No matching category for: ", data)

#     print("전처리가 완료되었습니다.")

# wellness_text_classification_data()
  
######################## wellness 대화 데이터셋 질문 답변 추출 후 저장#######################################

def wellness_dialog_for_autoregressive():
  root_path = "C:\\sqlite\\mysql\\code\\AI\\FINAL_project\\WellnessConversation-LanguageModel-master\\WellnessConversation-LanguageModel-master\\dataloader"
  wellness_file = root_path + "\\wellness_dialog_dataset.xlsx"
  wellness_answer_file = root_path + "\\wellness_dialog_answer.txt"
  wellness_question_file = root_path + "\\wellness_dialog_question.txt"
  wellness_autoregressive_file = root_path + "\\wellness_dialog_for_autoregressive.txt"

  with open(wellness_answer_file, 'r') as answ_file:
    answ_lines = answ_file.readlines()

  with open(wellness_question_file, 'r') as ques_file:
    ques_lines = ques_file.readlines()

  ans_dict = {ans_line.split('    ')[0]: ans_line.split('    ')[1] for ans_line in answ_lines}
  print("Answer Dictionary: ", ans_dict) 

  with open(wellness_autoregressive_file, 'w') as autoregressive_file:
    for line_num, line_data in enumerate(ques_lines):
      ques_data = line_data.split('    ')
      print(f"Question Data #{line_num}: {ques_data}")
      if ques_data[0] in ans_dict:
        autoregressive_file.write(ques_data[1][:-1] + "    " + ans_dict[ques_data[0]])


######################### 공감형 대화 셋 가공 후 저장#######################################

def tweet_data_for_autoregressive():
  root_path = "C:\\sqlite\\mysql\\code\\AI\\FINAL_project\\WellnessConversation-LanguageModel-master\\WellnessConversation-LanguageModel-master\\dataloader"
  file_path = root_path + "\\tweeter_dialog_data.txt"
  tweeter_autoregressive_file = root_path + "\\tweeter_dialog_for_autoregressive.txt"

  with open(file_path, 'r', encoding='utf-8') as data_file, open(tweeter_autoregressive_file, 'w', encoding='utf-8') as tweet_file:
    dialog = ''
    for line_data in data_file:
      if line_data == "\n" and dialog:
        dialog += "\n"
        tweet_file.write(dialog)
        print(dialog)
        dialog = ''
      elif line_data != "\n":
        dialog += "<s>" + line_data.strip() + "</s>"
  
######################### 공감형 대화 셋 텍스트 추출 #######################################

def extract_data_from_autoregressive_file():
    root_path = "C:\\sqlite\\mysql\\code\\AI\\FINAL_project\\WellnessConversation-LanguageModel-master\\WellnessConversation-LanguageModel-master\\dataloader"
    input_file_path = root_path + "\\tweeter_dialog_for_autoregressive.txt"
    output_file_path = root_path + "\\extracted_data.txt" 
    
    extracted_data = []

    with open(input_file_path, 'r', encoding='utf-8') as file:
        data = file.read()

    for section in re.finditer(r"<s>(.*?)</s>", data):
      try:
          corrected_section = section.group(1).replace("'", '"').replace("None", 'null').replace("True", 'true').replace("False", 'false')
          # 이 부분을 추가하여 문제의 문자열을 확인합니다.
          print("Corrected section for debugging:", corrected_section)
          tweet_obj = json.loads(corrected_section)

          text = tweet_obj.get('text')
          empathy = tweet_obj.get('listener_empathy')
          extracted_data.append({
                'text': text,
                'listener_empathy': empathy
            })
      except json.JSONDecodeError as e:
        # 에러가 발생한 구체적인 부분을 출력하여 문제를 진단합니다.
        print(f"JSON parsing error at section: {section.group(1)}")
        print(f"Error message: {e}")

    # 추출된 데이터를 새 파일에 쓰기
    with open(output_file_path, 'w', encoding='utf-8') as file:
        for item in extracted_data:
            file.write(json.dumps(item, ensure_ascii=False) + "\n")

######################## wellness 자기회귀 학습 가공 저장#######################################

# def seperate_wellness_data():
#   wellness_autoregressive_file = root_path+"/wellness_dialog_for_autoregressive.txt"
#   wellness_text_classification_file = root_path + "/wellness_dialog_for_text_classification.txt"
#   root_path = "C:\\sqlite\\mysql\\code\\AI\\FINAL_project\\WellnessConversation-LanguageModel-master\\WellnessConversation-LanguageModel-master\\dataloader"
#   file_path = root_path + "\\wellness_dialog_for_autoregressive.txt"
#   train_file_path = root_path + "\\wellness_dialog_for_autoregressive_train.txt"
#   test_file_path = root_path + "\\wellness_dialog_for_autoregressive_test.txt"

#   sperated_file = open(file_path, 'r')
#   train_file = open(train_file_path, 'w')
#   test_file = open(test_file_path, 'w')

#   sperated_file_lines = sperated_file.readlines()
#   ques_dict = {}
#   for line_num, line_data in enumerate(sperated_file_lines):
#     rand_num = random.randint(0, 10)
#     if rand_num < 10:
#       train_file.write(line_data)
#       print(f"Train Data: {line_data}")
#     else:
#       test_file.write(line_data)
#       print(f"Test Data: {line_data}")

#   sperated_file.close()
#   train_file.close()
#   test_file.close()


######################## tweeter 대화 셋 훈련, 테스트 분할 저장#######################################

def tweeter_autoregressive_data():
    root_path = "C:\\sqlite\\mysql\\code\\AI\\FINAL_project\\WellnessConversation-LanguageModel-master\\WellnessConversation-LanguageModel-master\\dataloader"
    tokenizer = get_kogpt2_tokenizer() 
    file_path = root_path + "\\tweeter_dialog_data.txt"
    tweeter_autoregressive_file = root_path + "\\tweeter_dialog_for_autoregressive.txt"

    with open(file_path, 'r', encoding='utf-8') as data_file, \
         open(tweeter_autoregressive_file, 'w', encoding='utf-8') as tweet_file:

      dialog = ''
      max_len = 0
      for line_num, line_data in enumerate(data_file.readlines()):
        if line_data == "\n" and dialog != '':
          dialog += "\n"
          tweet_file.write(dialog)
          print(dialog)
          dialog = ''
        elif line_data != "\n":
          tmp_data = dialog + "<s>" + line_data.strip() + "</s>" 
          encoded_len = len(tokenizer.encode(tmp_data))
          if encoded_len >= 1024:
            continue
          else:
            max_len = max(encoded_len, max_len)
            dialog = tmp_data
      print('max_token_length:', max_len)

####################### wellness 대화 셋 자연어 생성, 가공 후 토큰화 저장#######################################

# def wellness_autoregressive_data_with_token():
#     root_path = "C:\\sqlite\\mysql\\code\\AI\\FINAL_project\\WellnessConversation-LanguageModel-master\\WellnessConversation-LanguageModel-master\\dataloader"
#     wellness_answer_file = root_path + "\\wellness_dialog_answer.txt"
#     wellness_question_file = root_path + "\\wellness_dialog_question.txt"
#     wellness_autoregressive_file = root_path + "\\wellness_dialog_for_autoregressive_with_token.txt"

#     with open(wellness_answer_file, 'r') as answ_file, \
#          open(wellness_question_file, 'r') as ques_file, \
#          open(wellness_autoregressive_file, 'w', encoding='utf-8') as autoregressive_file:

#         answ_lines = answ_file.readlines()
#         ques_lines = ques_file.readlines()
        
#         for ques_line in ques_lines:
#             ques_data = ques_line.strip().split('    ')
#             for ans_line in answ_lines:
#                 ans_data = ans_line.strip().split('    ')
#                 if ques_data[0] == ans_data[0]:
#                     combined_data = "<s>" + ques_data[1] + "</s><s>" + ans_data[1] + "</s>\n"
#                     autoregressive_file.write(combined_data)
#                     print(combined_data)


######################## 합치기#######################################

def merge_data():
    root_path = "C:\\sqlite\\mysql\\code\\AI\\FINAL_project\\WellnessConversation-LanguageModel-master\\WellnessConversation-LanguageModel-master\\dataloader"

    chatbot_file = root_path + "\\chatbot_wellness_data.txt"
    wellness_file = root_path + "\\wellness_dialog_dataset.txt"
    total_data_file = root_path + "\\a_chatbot_wellness_data.txt"

    with open(chatbot_file, 'r', encoding='utf-8') as chatbot_f, \
         open(wellness_file, 'r') as wellness_f, \
         open(total_data_file, 'w', encoding='utf-8') as output_f:

        chatbot_lines = chatbot_f.readlines()
        for line in chatbot_lines: 
            # print(line, end='')
            output_f.write(line)

        wellness_lines = wellness_f.readlines()
        for line in wellness_lines:
            # print(line, end='')
            output_f.write(line)

if __name__ == "__main__":
    root_path = "C:\\sqlite\\mysql\\code\\AI\\FINAL_project\\WellnessConversation-LanguageModel-master\\WellnessConversation-LanguageModel-master\\dataloader"
    file_path = os.path.join(root_path, "chatbot_wellness_data.txt")
    o_path = os.path.join(root_path, "chatbot_wellness_data_for_autoregressive.txt")

    with open(file_path, 'r', encoding='utf-8') as i_file, open(o_path, 'w', encoding='utf-8') as o_file:
        i_lines = i_file.readlines()
        for i, data in enumerate(i_lines):
            tmp = data.strip().split(',')
            print(tmp)

            if len(tmp) >= 3:
                question = tmp[0]
                answer = tmp[1]
                label = tmp[2]
                o_file.write(f"<s>{question}</s><s>{answer}</s><s>{label}</s>\n")
            else:
                print(f"line {i + 1}은 리스트 길이가 짧아 건너뛰었습니다.")
  
# if __name__ == "__main__":
#     # 함수 호출 순서
#     tweet_dialog_dataset('C:\\sqlite\\mysql\\code\\AI\\FINAL_project\\WellnessConversation-LanguageModel-master\\WellnessConversation-LanguageModel-master\\dataloader\\train\\',
#      'C:\\sqlite\\mysql\\code\\AI\\FINAL_project\\WellnessConversation-LanguageModel-master\\WellnessConversation-LanguageModel-master\\dataloader\\tweeter_dialog_data.txt') # 트위터 대화셋 호출 후 저장
#     wellness_question_data() # Wellness 대화셋 질문부분을 추출 후 저장
#     wellness_answer_data() # Wellness 대화 데이터 셋 답변 부분 추출 후 저장
#     category_data() # Wellness 대화 데이터 셋 카테고리 확인, 저장
#     wellness_text_classification_data() # Wellness 대화 ,카테고리 맵핑 후 저장
#     wellness_dialog_for_autoregressive() # wellness 대화 데이터셋 질문 답변 추출 후 저장
#     tweet_data_for_autoregressive() # 트위터 대화 셋 가공 후 저장
#     extract_data_from_autoregressive_file() 
#     seperate_wellness_data() # Wellness 대화 데이터 셋 훈련, 테스트 세트 분리
#     tweeter_autoregressive_data() # wellness 자기회귀 학습 가공 저장
#     tweeter_autoregressive_data_with_token() # 트위터 대화 셋 자연어 생성, 가공 후 토큰화 저장
#     merge_data() # 합치기
