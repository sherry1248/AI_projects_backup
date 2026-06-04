import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl import Workbook
import re

# 주소록에서 주소만 가져오기
filePath = r'C:\\sqlite\\mysql\\code\\AI\\Mid_project\\data\\18e038dc4b5c4562 - 복사본.xlsx'
df_from_excel = pd.read_excel(filePath, engine='openpyxl')
# df_from_excel.columns = df_from_excel.loc[0].tolist()
# df_from_excel = df_from_excel.drop(index=list(range(0, 5)))

# print(df_from_excel)
# 오픈API 사용준비
url = 'http://api.vworld.kr/req/address?'
params = 'service=address&request=getcoord&version=2.0&crs=epsg:4326&refine=true&simple=false&format=json&type='
road_type = 'ROAD'  # 도로명 주소
road_type2 = 'PARCEL'  # 지번 주소
address = '&address='
keys = '&key='
primary_key = '683E2ED2-8E65-3440-BA7F-F66E58D3044D'

# # 좌표 얻어오는 함수


def request_geo(road):
    page = requests.get(url+params+road_type+address+road+keys+primary_key)
    json_data = page.json()
    if json_data['response']['status'] == 'OK':
        x = json_data['response']['result']['point']['x']
        y = json_data['response']['result']['point']['y']
        return x, y
    else:
        x = 0
        y = 0
        return x, y


# 얻어 온 좌표 엑셀로 저장
try:
    wb = load_workbook(
        r'C:\\sqlite\\mysql\\code\\AI\\Mid_project\\data\\좌표.xlsx', data_only=True)
    sheet = wb.active
except:
    wb = Workbook()
    sheet = wb.active

address_list = df_from_excel['시군구'].to_list()


for num, value in enumerate(address_list):
    # 주소에서 괄호 빼기(subtract)
    addr = re.sub(r'\([^)]*\)', '', value)
    print(addr)
    # 좌표 얻어오기
    x, y = request_geo(addr)
    # 엑셀에 써 넣기
    sheet.append([address_list[num], addr, x, y])

wb.save(r"C:\\sqlite\\mysql\\code\\AI\\Mid_project\\data\\좌표_2.xlsx")